import openpyxl

wb = openpyxl.load_workbook(r'C:\Claude_code\agent_changelog.xlsx')

today = '2026-05-25'

# 행 추가 데이터
# lead 탭: [날짜, 버전, 업데이트 내용, 업데이트 이유, 변경 요약, 트리거, 영향, 작성자]
# 나머지 탭: [날짜, 버전, 업데이트 내용, 업데이트 이유]

rows_to_add = {
    'lead': [
        today,
        'v6.70',
        '로컬 검증 게이트 신설 — 게임 위젯 PATCH 위임 전 game/qa 로컬 PASS 확인 의무',
        '2026-05-25 Post7 FSM 통합 라이브 사고 — 로컬 검증 없이 PATCH → 버그 라이브 배포',
        '로컬 검증 게이트 룰 신설 (F2)',
        'F2 룰화 작업',
        'lead, game, dev, qa, CLAUDE.md',
        'dev'
    ],
    'game': [
        today,
        'v3.45',
        '로컬 HTML 검증 절차 신설 — 코드 수정 후 로컬 wrapper.html 플레이 PASS 없이 납품 금지',
        '2026-05-25 Post7 FSM 통합 라이브 사고 재발 방지',
    ],
    'dev': [
        today,
        'v6.56',
        '라이브 PATCH 전 로컬 검증 PASS 확인 필수 룰 신설 — PASS 미첨부 시 PATCH 거부 의무',
        '2026-05-25 Post7 FSM 통합 라이브 사고 재발 방지',
    ],
    'qa': [
        today,
        'v8.58',
        '로컬 게임 플레이 검수 항목 신설 — 라이브 전용 검수 FAIL 처리, 로컬 체크리스트 의무화',
        '2026-05-25 Post7 FSM 통합 라이브 사고 재발 방지',
    ],
}

# CLAUDE.md 변경은 lead 탭에 추가 행으로 기록
rows_to_add_extra = {
    'lead': [
        today,
        'CLAUDE.md',
        '라이브 PATCH 전 로컬 HTML 검증 필수 섹션 신설 (★ 2026-05-25)',
        '2026-05-25 Post7 FSM 통합 라이브 사고 — CLAUDE.md 프로젝트 룰 반영',
        'CLAUDE.md 룰 신설 (F2)',
        'F2 룰화 작업',
        'CLAUDE.md',
        'dev'
    ]
}

for tab, row in rows_to_add.items():
    ws = wb[tab]
    ws.append(row)

# CLAUDE.md 추가 행
wb['lead'].append(rows_to_add_extra['lead'])

wb.save(r'C:\Claude_code\agent_changelog.xlsx')

with open(r'C:\Claude_code\automation\py_out_f2.txt', 'w', encoding='utf-8') as f:
    f.write('DONE\n')
    for tab, row in rows_to_add.items():
        ws = wb[tab]
        f.write(f'[{tab}] new rows={ws.max_row}\n')
    f.write(f"[lead] CLAUDE.md row added, total={wb['lead'].max_row}\n")

print('done')
